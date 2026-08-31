"""excel_exporter 双行表头导出单元测试。"""
import os
import tempfile
import unittest

import openpyxl

from excel_exporter import write_export_excel


class ExcelExporterTest(unittest.TestCase):
    """覆盖: 分类合并 / 双行表头 / 数据从第三行开始。"""

    def test_double_row_header_and_data(self):
        categories = [
            ("绑定单号", ["商家订单号", "平台订单号"]),
            ("收件人信息", ["收件人姓名"]),
        ]
        rows = [
            ["A1", "B1", "张三"],
            ["A2", "B2", "李四"],
        ]
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            write_export_excel(path, categories, rows)
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            self.assertEqual(ws.cell(1, 1).value, "绑定单号")
            self.assertEqual(ws.cell(1, 3).value, "收件人信息")
            self.assertEqual(ws.cell(2, 1).value, "商家订单号")
            self.assertEqual(ws.cell(2, 2).value, "平台订单号")
            self.assertEqual(ws.cell(2, 3).value, "收件人姓名")
            self.assertEqual(ws.cell(3, 1).value, "A1")
            self.assertEqual(ws.cell(3, 2).value, "B1")
            self.assertEqual(ws.cell(4, 3).value, "李四")
            self.assertEqual(ws.freeze_panes, "A3")
            self.assertGreaterEqual(len(ws.merged_cells.ranges), 1)


if __name__ == "__main__":
    unittest.main()
