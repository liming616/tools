"""
产地快打 — Excel 导出工具类

生成双行表头（第一行分类、第二行字段）的导出文件，
数据从第三行开始，业务层只负责传入分类结构与数据行。
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 导出表头 -> fields dict 字段路径映射；未列出的字段导出时留空
EXPORT_FIELD_MAP = {
    "寄件人姓名": "sender_name",
    "寄件人手机": "sender_phone",
    "寄件人座机": "sender_landline",
    "寄件人地址": "sender_address",
    "寄件人公司": "sender_company",
    "发货仓编码": "sender_warehouse_code",
    "收件人姓名": "name",
    "收件人手机": "phone",
    "收件人座机": "landline",
    "收件人地址": "full_address",
    "收件人公司": "company",
    "物品类型": "items_text",
    "时效产品": "delivery_product",
    "温层": "temperature_layer",
    "面单备注": "notes",
    "自定义信息": "raw",
}

# 表头分类默认配色（按类别顺序循环使用）
DEFAULT_HEADER_COLORS = [
    "D9E1F2", "FCE4D6", "E2EFDA", "FFF2CC", "DDEBF7",
    "F2DCDB", "E4DFEC", "D9F2E6", "FCE8D5", "DDEBED",
]

# 下拉选项隐藏 sheet 名称与数据验证行数
DROPDOWN_SHEET_NAME = "下拉选项"
DROPDOWN_VALIDATION_ROWS = 1000


def flatten_categories(categories) -> list:
    """将 [(分类名, [字段...], 颜色), ...] 展平为字段标题列表。"""
    headers = []
    for cat in categories:
        headers.extend(cat[1])
    return headers


def write_export_excel(file_path: str, categories: list, rows: list,
                       dropdown_options: dict = None) -> None:
    """生成双行表头 Excel 文件。

    Args:
        file_path: 保存路径（.xlsx）
        categories: [(分类名, [字段名...], 颜色), ...]，仅包含用户勾选的字段
        rows: 数据行，每行顺序与 categories 展平后的字段顺序一致
        dropdown_options: {表头: [选项...], ...}，为匹配的表头列生成下拉框
    """
    headers = flatten_categories(categories)

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产地快打导出"

    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="808080")
    header_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    col = 1
    for cat_idx, cat in enumerate(categories):
        category_name = cat[0]
        fields = cat[1]
        if not fields:
            continue
        fill_color = (
            str(cat[2]).strip()
            if len(cat) > 2 and str(cat[2]).strip()
            else DEFAULT_HEADER_COLORS[cat_idx % len(DEFAULT_HEADER_COLORS)]
        )
        header_fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        )
        start_col = col
        for field in fields:
            cell = ws.cell(row=2, column=col, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
            col += 1
        end_col = col - 1

        for c in range(start_col, end_col + 1):
            cat_cell = ws.cell(
                row=1, column=c,
                value=(category_name if c == start_col else None),
            )
            cat_cell.font = header_font
            cat_cell.fill = header_fill
            cat_cell.alignment = header_alignment
            cat_cell.border = header_border
        if end_col > start_col:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col,
            )

    # 下拉选项：写入隐藏 sheet，并按表头列挂载数据验证
    if dropdown_options:
        dropdown_fields = [
            (name, options)
            for name, options in dropdown_options.items()
            if isinstance(options, list) and options
        ]
        if dropdown_fields:
            dd_sheet = wb.create_sheet(DROPDOWN_SHEET_NAME)
            dd_sheet.sheet_state = "hidden"
            for col_idx, (field_name, options) in enumerate(dropdown_fields, start=1):
                dd_sheet.cell(row=1, column=col_idx, value=field_name)
                for row_idx, option in enumerate(options, start=2):
                    dd_sheet.cell(row=row_idx, column=col_idx, value=str(option))

            header_to_col = {
                header: col_idx
                for col_idx, header in enumerate(headers, start=1)
            }
            for col_idx, (field_name, options) in enumerate(dropdown_fields, start=1):
                target_col = header_to_col.get(field_name)
                if not target_col:
                    continue
                dd_letter = get_column_letter(col_idx)
                last_row = 1 + len(options)
                dv = DataValidation(
                    type="list",
                    formula1=(
                        f"'{DROPDOWN_SHEET_NAME}'!{dd_letter}$2:"
                        f"{dd_letter}${last_row}"
                    ),
                    allow_blank=True,
                    showDropDown=False,
                )
                ws.add_data_validation(dv)
                target_letter = get_column_letter(target_col)
                dv.add(
                    f"{target_letter}3:"
                    f"{target_letter}{2 + DROPDOWN_VALIDATION_ROWS}"
                )

    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(value))

    for col_idx, header in enumerate(headers, 1):
        width = max(14, min(36, len(str(header)) * 2 + 8))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"
    wb.save(file_path)
