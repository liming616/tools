"""
Excel 模版管理器 — 加载模版表头 + 字段映射 + 数据行构建

支持 .xlsx (openpyxl) 和 .xls (xlrd) 两种格式。
"""

import os
import logging
from typing import Optional

logger = logging.getLogger("moshidadan.template_manager")

# ======================== 默认表头 ========================

DEFAULT_HEADERS = ["姓名", "手机号", "地址"]

# ======================== 字段描述符映射表 ========================
# 格式: (field_path, 中文标签, [匹配关键词])
# field_path 支持点号分隔的嵌套路径（如 "addr.province"）

FIELD_DESCRIPTORS: list[tuple[str, str, list[str]]] = [
    # ---- 寄件人 / 发货方（来自预制信息，订单解析不填充）----
    ("sender_name",           "寄件人姓名", ["寄件人姓名", "发货人姓名", "发件人姓名"]),
    ("sender_phone",          "寄件人手机", ["寄件人手机", "发货人手机", "发件人手机", "发货人电话"]),
    ("sender_landline",       "寄件人座机", ["寄件人座机", "发货人座机", "发件人座机", "发货人固话"]),
    ("sender_address",        "寄件人地址", ["寄件人地址", "发货人地址", "发件人地址",
                                             "发货地址", "寄件地址"]),
    ("sender_company",        "寄件人公司", ["寄件人公司", "发货人公司", "发件人公司"]),
    ("sender_warehouse_code", "发货仓编码", ["发货仓编码", "仓库编码", "发货仓", "仓库编号"]),

    # ---- 收件人 / 收货方（来自订单解析）----
    ("name",           "姓名",     ["收件人姓名", "收货人", "收件人", "客户",
                                     "联系人", "名字", "下单人", "姓名"]),
    ("phone",          "手机号",   ["收件人手机", "收件人电话", "手机号", "联系电话",
                                     "联系方式", "电话", "手机", "号码"]),
    ("landline",       "座机",     ["收件人座机", "座机", "固话"]),
    ("province",       "省",       ["省份", "省"]),
    ("city",           "市",       ["城市", "市"]),
    ("district",       "区/县",    ["区县", "地区", "区域", "区", "县"]),
    ("township",       "街道/镇",  ["街道", "镇", "乡", "街道办"]),
    ("road",           "路/街",    ["路", "街", "大道", "巷", "弄"]),
    ("community",      "小区/村",  ["小区", "花园", "社区", "苑", "村"]),
    ("building",       "栋/楼",    ["号楼", "栋", "幢", "座"]),
    ("unit",           "单元",     ["单元", "门"]),
    ("room",           "室",       ["房号", "房间", "室"]),
    ("full_address",   "地址",     ["收件人地址", "收货地址", "收件地址",
                                     "邮寄地址", "送达地址", "详细地址", "具体地址",
                                     "地址"]),
    ("full_detail",    "详细地址",  ["详细地址", "具体地址"]),
    ("items_text",     "商品",     ["物品类型", "商品名称", "货品", "货物名称"]),
    ("delivery_product", "时效产品", ["时效产品", "产品时效", "时效"]),
    ("temperature_layer", "温层", ["温层", "温度层", "温区"]),
    ("notes",          "备注",     ["面单备注", "订单备注", "备注", "留言", "说明"]),
    ("raw",            "原始文本",  ["原文", "原始文本", "完整信息", "自定义信息"]),
    ("company",        "公司",     ["收件人公司", "公司", "单位", "企业"]),
    ("qq",             "QQ号",     ["QQ号", "QQ"]),
]


def load_template_headers(filepath: str) -> list[str]:
    """
    从 Excel 文件首行（或第二行，根据内容判断）读取表头。
    支持 .xlsx 和 .xls 格式。

    Returns:
        表头字符串列表（已去空白，过滤空值）
    Raises:
        ValueError: 文件无效或表头为空
    """
    if not os.path.exists(filepath):
        raise ValueError(f"文件不存在: {filepath}")

    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        headers = _load_xlsx_headers(filepath)
    elif ext == ".xls":
        headers = _load_xls_headers(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx / .xls")

    # 过滤空值并 strip
    headers = [h.strip() for h in headers if h and str(h).strip()]
    if not headers:
        raise ValueError("Excel 文件中未找到有效的表头行")

    logger.info("模版加载成功 | %s | %d 列表头", filepath, len(headers))
    return headers


def _load_xlsx_headers(filepath: str) -> list[str]:
    """从 .xlsx 文件读取表头。使用第二行（row 2），因为首行通常是合并的类别标题。"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # 读取前两行来判断哪行是实际的列标题
    row1 = [_safe_cell(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    row2 = [_safe_cell(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)]

    wb.close()

    # 如果 row2 有数据且 row1 看起来更像合并标题（更少非空值），使用 row2
    row1_count = sum(1 for v in row1 if v)
    row2_count = sum(1 for v in row2 if v)

    if row2_count > row1_count:
        return row2
    return row1


def _load_xls_headers(filepath: str) -> list[str]:
    """从 .xls 文件读取表头。"""
    import xlrd
    wb = xlrd.open_workbook(filepath, encoding_override='gb18030')
    ws = wb.sheet_by_index(0)

    if ws.nrows < 1:
        raise ValueError("Excel 文件为空")

    # 读取前两行，选择内容更多的作为实际表头
    row1 = [_safe_cell(ws.cell_value(0, c)) for c in range(ws.ncols)]
    row1_count = sum(1 for v in row1 if v)

    if ws.nrows >= 2:
        row2 = [_safe_cell(ws.cell_value(1, c)) for c in range(ws.ncols)]
        row2_count = sum(1 for v in row2 if v)
        if row2_count > row1_count:
            return row2

    return row1


def _safe_cell(value) -> str:
    """安全转换单元格值为字符串。"""
    if value is None:
        return ""
    # Excel 数值单元格读出来是 float，整数统一转 int 去掉 ".0"（如 15811111111.0 → 15811111111）
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    # 去掉 Excel 常见的前后空白和换行
    s = s.replace('\n', ' ').replace('\r', '')
    return s


# ======================== 预制信息读取 ========================


def load_prefill_rows(filepath: str, max_rows: int = 2) -> tuple[list[str], list[dict]]:
    """
    读取预制信息 Excel：表头 + 至多 max_rows 行数据。

    与模版读取共用同一套表头探测逻辑（首行/次行取非空更多者）。
    每一行数据作为一个「档案」，返回 {"表头": "值", ...} 的字典列表。

    Returns:
        (headers, rows)
    Raises:
        ValueError: 文件无效 / 无表头 / 无数据行
    """
    if not os.path.exists(filepath):
        raise ValueError(f"文件不存在: {filepath}")

    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        rows = _load_xlsx_rows(filepath)
    elif ext == ".xls":
        rows = _load_xls_rows(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx / .xls")

    headers, profiles = _extract_prefill(rows, max_rows)
    logger.info("预制信息读取成功 | %s | %d 列 | %d 档案", filepath, len(headers), len(profiles))
    return headers, profiles


def _load_xlsx_rows(filepath: str) -> list[list[str]]:
    """读取 .xlsx 全部行（含表头与数据）。"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    out = [[_safe_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return out


def _load_xls_rows(filepath: str) -> list[list[str]]:
    """读取 .xls 全部行（含表头与数据）。"""
    import xlrd
    wb = xlrd.open_workbook(filepath, encoding_override='gb18030')
    ws = wb.sheet_by_index(0)
    out = [[_safe_cell(ws.cell_value(r, c)) for c in range(ws.ncols)]
           for r in range(ws.nrows)]
    return out


def _extract_prefill(rows: list[list[str]], max_rows: int,
                     allow_empty: bool = False) -> tuple[list[str], list[dict]]:
    """从原始行中提取表头与至多 max_rows 个档案。

    allow_empty=True 时，无数据行不报错，返回空档案列表（用于「仅模版」场景）。
    """
    if not rows:
        raise ValueError("Excel 文件为空")

    # 表头行：首行/次行取非空更多者
    header_idx = 0
    if len(rows) >= 2:
        c0 = sum(1 for v in rows[0] if v)
        c1 = sum(1 for v in rows[1] if v)
        if c1 > c0:
            header_idx = 1

    header_row = rows[header_idx]
    cols = [(i, h) for i, h in enumerate(header_row) if h]
    if not cols:
        raise ValueError("预制信息Excel中未找到有效的表头")

    headers = [h for _, h in cols]
    col_idx_list = [i for i, _ in cols]

    profiles: list[dict] = []
    for r in rows[header_idx + 1:]:
        if len(profiles) >= max_rows:
            break
        values = {}
        for ci, h in zip(col_idx_list, headers):
            values[h] = r[ci] if ci < len(r) else ""
        if any(values.values()):
            profiles.append(values)

    if not profiles and not allow_empty:
        raise ValueError("预制信息Excel中未找到数据行")

    return headers, profiles


def load_template_and_prefill(filepath: str, max_prefill_rows: int = 2) -> tuple[list[str], list[dict]]:
    """
    一次性读取模版 Excel：表头（作为模版列）+ 至多 max_prefill_rows 行预制数据。

    与 load_template_headers 共用同一套表头探测逻辑；数据行作为「预制信息档案」返回。
    无数据行时返回空档案列表（不报错），供「仅模版」场景使用。

    Returns:
        (headers, profiles) — profiles 为 [{"表头": "值", ...}, ...]
    Raises:
        ValueError: 文件无效 / 无表头
    """
    if not os.path.exists(filepath):
        raise ValueError(f"文件不存在: {filepath}")

    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        rows = _load_xlsx_rows(filepath)
    elif ext == ".xls":
        rows = _load_xls_rows(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx / .xls")

    headers, profiles = _extract_prefill(rows, max_prefill_rows, allow_empty=True)
    logger.info("模版+预制信息读取成功 | %s | %d 列 | %d 档案",
                filepath, len(headers), len(profiles))
    return headers, profiles


# ======================== 字段映射 ========================


def map_headers_to_fields(headers: list[str]) -> list[Optional[str]]:
    """
    将表头列表映射为 field_path 列表。

    匹配策略：
    1. 精确匹配：表头与某关键词完全一致
    2. 包含匹配：表头中包含某关键词
    3. 多个匹配时，选关键词最长的（更具体）

    未匹配的列返回 None。

    Args:
        headers: Excel 表头文本列表

    Returns:
        field_path 列表（长度与 headers 相同）
    """
    mapped: list[Optional[str]] = []

    for header in headers:
        header_clean = header.strip()
        best_field = None
        best_score = 0  # 分数越高匹配越好

        for field_path, _label, keywords in FIELD_DESCRIPTORS:
            for kw in keywords:
                score = 0
                # 精确匹配：最高优先级
                if header_clean == kw:
                    score = len(kw) * 100
                # 包含匹配
                elif kw in header_clean:
                    # 短关键词（<=2字）在长表头中出现假匹配概率高，给低分
                    if len(kw) <= 2 and len(header_clean) > 4:
                        score = 0  # 忽略
                    else:
                        # 分数 = 关键词长度 / 表头长度（越精确分越高）
                        score = len(kw) * 10 * len(kw) // len(header_clean)

                if score > best_score:
                    best_score = score
                    best_field = field_path

        mapped.append(best_field if best_score > 0 else None)

    matched_count = sum(1 for f in mapped if f is not None)
    logger.debug("表头映射完成 | %d/%d 列已匹配", matched_count, len(headers))
    return mapped


def get_field_label(field_path: str) -> str:
    """获取字段的中文标签。"""
    for fp, label, _keywords in FIELD_DESCRIPTORS:
        if fp == field_path:
            return label
    return field_path


# ======================== 数据行构建 ========================


def build_row_tuple(fields: dict, mapped_fields: list[Optional[str]]) -> tuple:
    """
    根据映射关系，从 fields dict 构建 Treeview 显示行。

    Args:
        fields: 解析后的字段字典
        mapped_fields: map_headers_to_fields 的返回值

    Returns:
        字符串元组（长度 = len(mapped_fields)）
    """
    values = []
    for field_path in mapped_fields:
        if field_path is None:
            values.append("")
            continue

        value = _resolve_field(fields, field_path)
        values.append(str(value) if value else "")

    return tuple(values)


def _resolve_field(fields: dict, field_path: str) -> str:
    """从 fields dict 中按路径取值。"""
    # items_text 特殊处理：直接存储识别到的物品原文，不做 name×qty 加工
    if field_path == "items_text":
        items = fields.get("items", [])
        if not items:
            return fields.get("items_text", "")
        return "; ".join(
            str(it.get("raw") or it.get("name", ""))
            for it in items
            if (it.get("raw") or it.get("name"))
        )

    # raw 特殊处理
    if field_path == "raw":
        return fields.get("raw", "")

    # 普通字段直接取值
    return fields.get(field_path, "")


def empty_fields_dict() -> dict:
    """返回一个所有已知字段为空的字典，作为 fields 模板。"""
    fields = {
        "name": "", "phone": "", "landline": "", "province": "", "city": "",
        "district": "", "township": "", "road": "", "community": "",
        "building": "", "unit": "", "room": "",
        "full_address": "", "full_detail": "",
        "items": [], "items_text": "", "delivery_product": "",
        "temperature_layer": "", "notes": "", "raw": "",
        "company": "", "qq": "",
        # 寄件人 / 发货方（来自预制信息）
        "sender_name": "", "sender_phone": "", "sender_landline": "", "sender_address": "",
        "sender_company": "", "sender_warehouse_code": "",
        # 地址解析特有字段（备用）
        "development_zone": "", "landmark": "",
    }
    return fields


def compute_column_width(field_path: Optional[str]) -> int:
    """根据字段类型返回建议列宽（px）。"""
    if field_path is None:
        return 70

    wide_fields = {"full_address", "full_detail", "items_text", "delivery_product",
                   "notes", "raw", "sender_address"}
    medium_fields = {"name", "phone", "landline", "province", "city", "district",
                     "township", "road", "community", "landmark",
                     "building", "unit", "room", "company",
                     "sender_name", "sender_phone", "sender_landline", "sender_company",
                     "sender_warehouse_code", "temperature_layer"}

    if field_path in wide_fields:
        return 250
    elif field_path in medium_fields:
        return 100
    else:
        return 80
